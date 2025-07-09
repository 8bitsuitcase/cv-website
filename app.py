from flask import Flask, request, jsonify, render_template
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen import canvas
import os

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'data/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'doc', 'docx'}

EXCEL_PATH = 'data/all_cvs.xlsx'
PDF_FOLDER = 'data/pdfs'

os.makedirs(PDF_FOLDER, exist_ok=True)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

FIELDS = [
    'name', 'age', 'nationality', 'current_salary', 'expected_salary',
    'email', 'phone', 'job_title', 'specialization', 'uploaded_file'
]

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def save_to_excel(data):
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(FIELDS)
    else:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        # Add header only if file is empty
        if ws.max_row == 0:
            ws.append(FIELDS)

    # Ensure the order of data matches the header
    row = [data.get(field, '') for field in FIELDS]
    ws.append(row)

    # Adjust column width for neatness
    for i, field in enumerate(FIELDS, start=1):
        column_width = max(len(str(field)), 15)
        ws.column_dimensions[get_column_letter(i)].width = column_width

    wb.save(EXCEL_PATH)

def generate_pdf(data):
    filename = f"{data['name'].replace(' ', '_')}_summary.pdf"
    path = os.path.join(PDF_FOLDER, filename)
    c = canvas.Canvas(path, pagesize=LETTER)
    text = c.beginText(40, 750)
    text.setFont("Helvetica", 12)
    for key, value in data.items():
        text.textLine(f"{key}: {value}")
    c.drawText(text)
    c.save()
    return filename

@app.route('/')
def home():
    return render_template('new_index.html')

@app.route('/submit-cv', methods=['POST'])
def submit_cv():
    form_data = request.form.to_dict()
    file = request.files.get('cv_file')

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        form_data['uploaded_file'] = filename
    else:
        form_data['uploaded_file'] = 'None'

    save_to_excel(form_data)
    pdf_filename = generate_pdf(form_data)

    return jsonify({'message': f"CV submitted successfully. PDF: {pdf_filename}."})

if __name__ == '__main__':
    app.run(debug=True)